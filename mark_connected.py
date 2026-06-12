"""Subagent 2/3: Match a JSON list of LinkedIn connections against the Excel
contacts and (optionally) mark the matches as 'Inviato'='y'.

This script does NOT open a browser. It reads:
  - connections_list.json  (output of scrape_connections.py)
  - Elenco linkedin (1).xlsx

It writes:
  - connections_match_report.csv (always)
  - Elenco linkedin (1).xlsx (only if --apply is set)
"""

import os
import sys
import argparse
import json
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "Elenco linkedin (1).xlsx"
SHEET_NAME = "Contatti"
DEFAULT_JSON = SCRIPT_DIR / "connections_list.json"
DEFAULT_REPORT = SCRIPT_DIR / "connections_match_report.csv"


def normalize_name(name: str) -> str:
    if not name:
        return ''
    n = name.strip().lower()
    n = (n.replace('à', 'a').replace('è', 'e').replace('é', 'e')
           .replace('ì', 'i').replace('ò', 'o').replace('ù', 'u'))
    n = re.sub(r'\s+', ' ', n)
    return n


def name_tokens(name: str) -> list[str]:
    return [t for t in normalize_name(name).split() if len(t) >= 2]


def matches(a_tokens: list[str], b_tokens: list[str]) -> bool:
    if not a_tokens or not b_tokens:
        return False
    sa, sb = set(a_tokens), set(b_tokens)
    if sa == sb:
        return True
    smaller, larger = (sa, sb) if len(sa) <= len(sb) else (sb, sa)
    return smaller.issubset(larger)


def main():
    parser = argparse.ArgumentParser(description="Subagent 2: Match connections JSON against Excel contacts.")
    parser.add_argument("--xlsx", default=str(XLSX_PATH), help="Excel file path.")
    parser.add_argument("--json", default=str(DEFAULT_JSON), help="Connections JSON file path.")
    parser.add_argument("--report", default=str(DEFAULT_REPORT), help="Output CSV report path.")
    parser.add_argument("--apply", action="store_true", help="Apply marks to the Excel file. Default: dry-run.")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    json_path = Path(args.json)
    report_path = Path(args.report)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"Connections JSON not found: {json_path}")

    names = json.loads(json_path.read_text(encoding='utf-8'))
    print(f"Loaded {len(names)} connection names from {json_path}")

    conn_tokens = [name_tokens(n) for n in names]
    conn_norm_set = {normalize_name(n) for n in names}

    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]

    def _ensure_header(col_idx, name):
        current = ws.cell(row=1, column=col_idx).value
        if current is None or str(current).strip() == '':
            ws.cell(row=1, column=col_idx, value=name)
    _ensure_header(5, 'Inviato')
    _ensure_header(6, 'Inviato il')

    matches_found: list[dict] = []
    y_rows: list[tuple[int, str]] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        contattare = ws.cell(row=r, column=4).value
        inviato = ws.cell(row=r, column=5).value
        if name is None:
            continue
        contattare_str = str(contattare).strip().lower() if contattare is not None else ''
        inviato_str = str(inviato).strip().lower() if inviato is not None else ''
        if contattare_str != 'y' or inviato_str == 'y':
            continue
        row_name = str(name).strip()
        y_rows.append((r, row_name))
        row_tokens = name_tokens(row_name)
        row_norm = normalize_name(row_name)
        is_match = False
        matched_name = None
        for tokens, conn_name in zip(conn_tokens, names):
            if matches(row_tokens, tokens):
                is_match = True
                matched_name = conn_name
                break
        if is_match:
            matches_found.append({
                'row': r,
                'excel_name': row_name,
                'linkedin_name': matched_name,
                'excel_normalized': row_norm,
                'linkedin_normalized': normalize_name(matched_name),
            })

    with open(report_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['row', 'excel_name', 'linkedin_name', 'excel_normalized', 'linkedin_normalized'])
        writer.writeheader()
        for m in matches_found:
            writer.writerow(m)
    print(f"Wrote {len(matches_found)} matches to {report_path}")

    print("\n=== MATCH SUMMARY ===")
    for m in matches_found:
        print(f"  Row {m['row']:4d}: '{m['excel_name']}' <-> '{m['linkedin_name']}'")

    matched_rows = {m['row'] for m in matches_found}
    unmatched = [(r, n) for r, n in y_rows if r not in matched_rows]
    print(f"\nTotal y contacts to process: {len(y_rows)}")
    print(f"  Already connected (will be marked): {len(matches_found)}")
    print(f"  To be sent to: {len(unmatched)}")
    if unmatched and len(unmatched) <= 30:
        print("\n  Unmatched y contacts:")
        for r, n in unmatched:
            print(f"    Row {r}: {n}")

    if args.apply:
        for m in matches_found:
            ws.cell(row=m['row'], column=5, value='y')
            ws.cell(row=m['row'], column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        wb.save(xlsx_path)
        print(f"\nMarked {len(matches_found)} contacts as already connected in {xlsx_path}.")
    else:
        print("\n*** DRY-RUN mode: Excel NOT modified. Re-run with --apply to apply. ***")


if __name__ == "__main__":
    main()
