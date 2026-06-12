"""Generate a complete report of all 187 y contacts from the Excel.

Reads the Excel and parse all run logs to determine the outcome for each contact:
- already_connected: script saw Message/Pending, or subagent 2 scrape match
- sent: script clicked Send (only v2 confirms with 'submitted' / modal close)
- send_unconfirmed: clicked Connect, modal opened, but no Send button found
- no_connect: rate limit / Follow only / no matching card
- no_results: search returned 0
- not_processed: row was 'y' but never seen in any log

Output: contact_report.csv + a summary printed to stdout (utf-8 safe).
"""

import re
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

# Force utf-8 stdout
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "Elenco linkedin (1).xlsx"
SHEET_NAME = "Contatti"
REPORT_PATH = SCRIPT_DIR / "contact_report.csv"
LOG_FILES = [
    SCRIPT_DIR / "run_full.log",
    SCRIPT_DIR / "test_run.log",
    SCRIPT_DIR / "send_run.log",
    SCRIPT_DIR / "send_v2.log",
    SCRIPT_DIR / "mark_run.log",
    SCRIPT_DIR / "scrape_run.log",
]

# Regex: "[N/M] Row R: Name"
ROW_HEADER_RE = re.compile(r'^\[(\d+)/(\d+)\] Row (\d+): (.*)$')
ROW_SCRAPE_RE = re.compile(r"Row (\d+): '([^']+)' <-> '([^']+)'")
SENT_CONFIRMED_RE = re.compile(
    r'(Invite appears to have been submitted|Invitation sent|invite.*submitted|request sent|'
    r'clicked via JS.*Send|clicked Send via JS|Send button clicked)'
)


def parse_log(path: Path):
    """Parse one log file. Returns dict: row_idx -> outcome."""
    if not path.exists():
        return {}
    try:
        text = path.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return {}

    outcomes = {}
    # v2-style log: one line "[N/M] Row R: Name\n  connected (Message)"
    # Build a list of (row, name, body) chunks.
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        m = ROW_HEADER_RE.match(lines[i])
        if m:
            row_idx = int(m.group(3))
            name = m.group(4).strip()
            # Collect body until next header or blank-then-header
            body_lines = []
            j = i + 1
            while j < len(lines):
                if ROW_HEADER_RE.match(lines[j]):
                    break
                body_lines.append(lines[j])
                j += 1
            body = '\n'.join(body_lines)

            outcome = None
            # Sent confirmed (very rare in v1, common in v2 with explicit success msg)
            if SENT_CONFIRMED_RE.search(body):
                outcome = 'sent'
            # already_connected patterns
            elif (
                'Marked as already connected' in body
                or 'connected (Message)' in body
                or 'already pending' in body
                or 'Person is already a connection' in body
                or 'Person found but no Connect action' in body
            ):
                outcome = 'already_connected'
            elif (
                'No Connect action on matching card' in body
                or 'Follow only' in body
                or 'follow only' in body
            ):
                outcome = 'no_connect'
            elif (
                'No Send button or confirmation found' in body
                or 'Send button or modal not found' in body
            ):
                outcome = 'send_unconfirmed'
            elif 'no_results' in body or 'No results' in body or 'no match' in body:
                outcome = 'no_results'
            else:
                # If 'Clicking Connect' was the last action, this is a click attempt
                if 'Clicking Connect' in body:
                    outcome = 'send_unconfirmed'
                else:
                    outcome = 'unknown'

            # Only overwrite if the new outcome is "stronger" (more specific).
            # Priority: sent > already_connected > send_unconfirmed > no_connect > no_results > unknown
            priority = {
                'sent': 5, 'already_connected': 4, 'send_unconfirmed': 3,
                'no_connect': 2, 'no_results': 2, 'unknown': 0, 'error': 1,
            }
            existing = outcomes.get(row_idx, ('unknown', -1))
            if priority.get(outcome, 0) > existing[1]:
                outcomes[row_idx] = (outcome, priority[outcome])
            i = j
        else:
            # Mark run log lines: "Row N: 'X' <-> 'Y'"
            sm = ROW_SCRAPE_RE.search(lines[i])
            if sm:
                row_idx = int(sm.group(1))
                # If subagent 2 marked this row, it's "already_connected (from scrape)"
                existing = outcomes.get(row_idx, ('unknown', -1))
                if existing[1] < 4:
                    outcomes[row_idx] = ('already_connected', 4)
            i += 1
    return {k: v[0] for k, v in outcomes.items()}


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    # Aggregate outcomes from all log files
    outcomes = {}
    for log in LOG_FILES:
        for row, outcome in parse_log(log).items():
            priority = {
                'sent': 5, 'already_connected': 4, 'send_unconfirmed': 3,
                'no_connect': 2, 'no_results': 2, 'unknown': 0, 'error': 1,
            }
            if priority.get(outcome, 0) > priority.get(outcomes.get(row, 'unknown'), 0):
                outcomes[row] = outcome

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        contattare = ws.cell(row=r, column=4).value
        inviato = ws.cell(row=r, column=5).value
        inviato_il = ws.cell(row=r, column=6).value
        if name is None:
            continue
        if str(contattare or '').strip().lower() != 'y':
            continue
        rows.append({
            'row': r,
            'name': str(name).strip(),
            'inviato': str(inviato or '').strip().lower(),
            'inviato_il': str(inviato_il or '').strip(),
            'outcome': outcomes.get(r, 'not_processed'),
        })

    # Write CSV
    with open(REPORT_PATH, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['row', 'name', 'inviato', 'inviato_il', 'outcome'])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    # Categorize
    by_outcome = {}
    for r in rows:
        by_outcome.setdefault(r['outcome'], []).append(r)

    print(f"Total y contacts: {len(rows)}")
    print(f"Report written to: {REPORT_PATH}")
    print()
    print("=== BY OUTCOME ===")
    for outcome, lst in sorted(by_outcome.items(), key=lambda x: -len(x[1])):
        print(f"  {outcome:25s} {len(lst):3d}")

    print()
    print("=== INVIATO FLAG ===")
    inviato_y = sum(1 for r in rows if r['inviato'] == 'y')
    inviato_n = sum(1 for r in rows if r['inviato'] != 'y')
    print(f"  Inviato='y'   : {inviato_y}")
    print(f"  Inviato!=y    : {inviato_n}")

    print()
    print("=== NOT YET CONNECTED ===")
    print("(inviti non ancora accettati, ovvero 'sent' / 'send_unconfirmed' / 'no_connect' / 'no_results' / 'not_processed')")
    not_connected_outcomes = {'sent', 'send_unconfirmed', 'no_connect', 'no_results', 'not_processed'}
    not_connected = [r for r in rows if r['outcome'] in not_connected_outcomes]
    print(f"  Count: {len(not_connected)}")
    print()
    for r in not_connected:
        print(f"  Row {r['row']:4d}: {r['name']:45s} [{r['outcome']:18s}] [{r['inviato'] or '-'}] {r['inviato_il']}")


if __name__ == "__main__":
    main()
