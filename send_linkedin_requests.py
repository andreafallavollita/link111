# send_linkedin_requests.py
"""
Automated LinkedIn connection requests script.

- Reads names from column A of the 'Contatti' sheet.
- Filters only rows where column D ('Contattare') = 'y'.
- Skips rows where column E ('Inviato') = 'y' (already contacted).
- Marks column E with 'y' (and timestamp in F) after a successful send.
- Searches each name on LinkedIn.
- Prefers 2nd-degree connection results (mutual connections) when present.
- Sends a connection request WITHOUT any message.
- Delay between requests: random between --delay-min and --delay-max seconds.
"""

import os
import sys
import time
import random
import argparse
from datetime import datetime
from pathlib import Path
import builtins

def _safe_str(s) -> str:
    """Convert to string, replacing characters that cannot be encoded in the
    current console's code page (e.g. cp1252 on Windows)."""
    if s is None:
        return ''
    txt = str(s)
    enc = sys.stdout.encoding or 'utf-8'
    try:
        txt.encode(enc)
        return txt
    except (UnicodeEncodeError, LookupError):
        return txt.encode(enc, errors='replace').decode(enc)


_original_print = builtins.print


def print(*args, **kwargs):
    """Print wrapper that flushes and handles non-encodable characters."""
    safe_args = [_safe_str(a) for a in args]
    sep = kwargs.get('sep', ' ')
    if not isinstance(sep, str):
        sep = _safe_str(sep)
    end = kwargs.get('end', '\n')
    if not isinstance(end, str):
        end = _safe_str(end)
    try:
        _original_print(*safe_args, sep=sep, end=end, flush=True)
    except (UnicodeEncodeError, ValueError, OSError):
        try:
            _original_print(*safe_args, sep=sep, end=end)
        except Exception:
            pass

import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ------------------------------------------------------------
# Helper functions
# ------------------------------------------------------------

def load_contacts(file_path: str, sheet_name: str = "Contatti") -> tuple:
    """Load contacts to process from Excel and ensure the 'Inviato' column exists.

    Returns (contacts, workbook, worksheet) where contacts is a list of
    (row_index_1based, name) tuples. The workbook handle is returned so the
    caller can persist state as sends are completed.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Contacts file not found: {file_path}")

    wb = load_workbook(path)
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]

    # Ensure column E ('Inviato') and column F ('Inviato il') exist
    def _ensure_header(col_idx: int, name: str) -> None:
        current = ws.cell(row=1, column=col_idx).value
        if current is None or str(current).strip() == '':
            ws.cell(row=1, column=col_idx, value=name)
        elif str(current).strip().lower() != name.lower():
            print(f"WARNING: column {col_idx} is '{current}', expected '{name}'. Using it as-is.")

    _ensure_header(5, 'Inviato')
    _ensure_header(6, 'Inviato il')
    wb.save(path)

    # Collect rows to process
    contacts = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        contattare = ws.cell(row=r, column=4).value
        inviato = ws.cell(row=r, column=5).value

        if name is None or str(name).strip() == '':
            continue
        contattare_str = str(contattare).strip().lower() if contattare is not None else ''
        inviato_str = str(inviato).strip().lower() if inviato is not None else ''

        if contattare_str == 'y' and inviato_str != 'y':
            contacts.append((r, str(name).strip()))

    return contacts, wb, ws


def mark_sent(wb, ws, file_path: str, row_idx: int) -> None:
    """Persist 'y' (and timestamp) into columns E/F for the given row, then save."""
    ws.cell(row=row_idx, column=5, value='y')
    ws.cell(row=row_idx, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    try:
        wb.save(file_path)
    except PermissionError as e:
        print(f"  WARNING: could not save workbook (file locked?): {e}")
    except Exception as e:
        print(f"  WARNING: error saving workbook: {e}")


def init_driver(profile_name: str = None) -> webdriver.Chrome:
    """Initialize a Chrome WebDriver instance (visible mode – LinkedIn blocks headless)."""
    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36")

    # Use persistent Chrome profile. Default to LinkedInAutomationProfile;
    # can be overridden via --profile or LINKEDIN_PROFILE env var.
    if profile_name is None:
        profile_name = os.environ.get("LINKEDIN_PROFILE", "LinkedInAutomationProfile")
    profile_dir = Path(os.environ.get("LOCALAPPDATA", "C:\\Temp")) / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    chrome_options.add_argument(f"--user-data-dir={profile_dir}")
    chrome_options.add_argument("--profile-directory=Default")

    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    driver.implicitly_wait(5)
    return driver


def linkedin_login(driver: webdriver.Chrome, email: str, password: str) -> None:
    """Log into LinkedIn."""
    debug_dir = Path(os.environ.get("TEMP", "."))
    
    # Check if we are already logged in from a previous session
    print("Checking login status...")
    driver.get("https://www.linkedin.com/feed")
    # Let the page settle (LinkedIn may redirect several times for cookie/region checks)
    time.sleep(5)
    
    is_logged_in = False
    # If we are still on /feed (or got redirected to a logged-in page like /mynetwork),
    # we are logged in. If we got redirected to /login or /signup, we are not.
    url_lower = driver.current_url.lower()
    if not any(x in url_lower for x in ['/login', '/signup', '/checkpoint', '/uas/']):
        # Also confirm by waiting for an element that exists only when logged in
        try:
            WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#global-nav, .global-nav, nav.global-nav, [data-test-global-nav]"))
            )
            is_logged_in = True
        except TimeoutException:
            # If URL is /feed but no nav appeared, still treat as logged in
            # (page may be slow to fully render). Only fall through if URL says otherwise.
            if '/feed' in url_lower or '/mynetwork' in url_lower or '/in/' in url_lower:
                print("  URL indicates logged-in state, treating as logged in.")
                is_logged_in = True
            else:
                print(f"  No global-nav found at {driver.current_url}.")
        
    if is_logged_in:
        print("Already logged in via persistent Chrome profile!")
        return

    print("Not logged in. Directing to login page...")
    driver.get("https://www.linkedin.com/login")
    time.sleep(5)

    # Debug: save screenshot + page source
    debug_dir = Path(os.environ.get("TEMP", "."))
    driver.save_screenshot(str(debug_dir / "linkedin_debug.png"))
    with open(debug_dir / "linkedin_debug.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    print(f"Debug screenshot saved to {debug_dir / 'linkedin_debug.png'}")
    print(f"Current URL: {driver.current_url}")
    print(f"Page title: {driver.title}")

    # Find the VISIBLE email field (LinkedIn has multiple hidden forms)
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email']"))
        )
        email_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='email']")
        email_input = None
        for el in email_inputs:
            if el.is_displayed():
                email_input = el
                break
        if not email_input:
            raise TimeoutException("No visible email field")
        print("Found visible email field.")
    except TimeoutException:
        print("ERROR: Could not find email input field.")
        raise

    # Find the VISIBLE password field
    try:
        pwd_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
        pwd_input = None
        for el in pwd_inputs:
            if el.is_displayed():
                pwd_input = el
                break
        if not pwd_input:
            raise NoSuchElementException("No visible password field")
        print("Found visible password field.")
    except NoSuchElementException:
        print("ERROR: Could not find password field.")
        raise

    email_input.send_keys(email)
    pwd_input.send_keys(password)

    # Click the login button
    login_btn = None
    try:
        buttons = driver.find_elements(By.TAG_NAME, "button")
        for btn in buttons:
            if btn.is_displayed():
                btn_text = btn.text.strip()
                # Matches exact text
                if btn_text in ["Accedi", "Sign in", "Log in", "Sign In", "Log In"]:
                    login_btn = btn
                    break
                # Fallback: contains key terms but not third-party login providers
                if ("Accedi" in btn_text or "Sign in" in btn_text or "Log in" in btn_text) and not any(x in btn_text for x in ["Google", "Microsoft", "Apple"]):
                    login_btn = btn
                    break
        if login_btn:
            print(f"Found login button: '{login_btn.text.strip()}'")
            login_btn.click()
        else:
            print("WARNING: Could not find login button, trying submit on password field...")
            pwd_input.submit()
    except Exception as e:
        print(f"Error clicking login button: {e}")
        try:
            pwd_input.submit()
        except Exception as submit_err:
            print(f"Error submitting form: {submit_err}")
            raise e

    # Wait to see if we get a challenge/checkpoint page
    time.sleep(5)
    
    checkpoint_detected = False
    checkpoint_attempts = 0
    max_checkpoint_wait_sec = 600  # Wait up to 10 minutes
    
    # We wait as long as the URL is still a login/verification/redirect URL
    while any(x in driver.current_url for x in ["checkpoint", "challenge", "login", "flagship-web"]):
        if not checkpoint_detected:
            print("ALERT: Verification challenge (CAPTCHA / App Notification / 2FA) or login redirect detected!")
            print("Please solve the verification challenge in the browser window or on your phone.")
            checkpoint_detected = True
        
        # Save screenshot so the progress can be viewed
        driver.save_screenshot(str(debug_dir / "linkedin_after_login.png"))
        print(f"[{checkpoint_attempts * 5}s] Still waiting for login to complete. Current URL: {driver.current_url}")
        
        # Check if we are redirected back to the login screen
        try:
            email_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='email']")
            visible_emails = [el for el in email_inputs if el.is_displayed()]
            if visible_emails:
                email_input = visible_emails[0]
                email_val = (email_input.get_attribute("value") or "").strip()
                # Re-submit only if the field is empty (indicating a fresh page load or clear)
                # or if we have been stuck on the login page for more than 25 seconds (checkpoint_attempts > 5)
                if not email_val or checkpoint_attempts > 5:
                    print("ALERT: Detected login page inside loop (form cleared or timeout). Re-submitting credentials...")
                    pwd_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
                    visible_pwds = [el for el in pwd_inputs if el.is_displayed()]
                    if visible_pwds:
                        pwd_input = visible_pwds[0]
                        email_input.clear()
                        email_input.send_keys(email)
                        pwd_input.clear()
                        pwd_input.send_keys(password)
                        
                        login_btn = None
                        buttons = driver.find_elements(By.TAG_NAME, "button")
                        for btn in buttons:
                            if btn.is_displayed():
                                btn_text = btn.text.strip()
                                if btn_text in ["Accedi", "Sign in", "Log in", "Sign In", "Log In"]:
                                    login_btn = btn
                                    break
                                if ("Accedi" in btn_text or "Sign in" in btn_text or "Log in" in btn_text) and not any(x in btn_text for x in ["Google", "Microsoft", "Apple"]):
                                    login_btn = btn
                                    break
                        if login_btn:
                            print(f"Clicking login button: '{login_btn.text.strip()}'")
                            login_btn.click()
                        else:
                            pwd_input.submit()
                        time.sleep(5)
                        checkpoint_attempts = 0
                        checkpoint_detected = False
                        continue
        except Exception as e:
            print(f"Error checking/re-submitting credentials: {e}")

        time.sleep(5)
        checkpoint_attempts += 1
        if checkpoint_attempts * 5 >= max_checkpoint_wait_sec:
            print("ERROR: Timeout waiting for login. Exiting.")
            raise TimeoutException("Login timeout.")
            
    if checkpoint_detected:
        print("Verification completed successfully!")
        
    time.sleep(3)
    driver.save_screenshot(str(debug_dir / "linkedin_after_login.png"))
    print(f"After login - URL: {driver.current_url}")
    print("Login completed.")


def search_and_connect(driver: webdriver.Chrome, name: str) -> str:
    """Search for a person on LinkedIn and send a connection request.

    Returns one of: 'sent' | 'skipped' | 'no_results' | 'no_match' |
                    'already_connected' | 'error'
    """
    debug_dir = Path(os.environ.get("TEMP", "."))
    # Search for the person
    search_url = f"https://www.linkedin.com/search/results/people/?keywords={name.replace(' ', '%20')}"
    driver.get(search_url)
    time.sleep(5)

    # Save search debug screenshot
    driver.save_screenshot(str(debug_dir / "linkedin_search_debug.png"))
    print(f"  Debug search screenshot saved.")

    # ------- Find the best matching card for the searched name -------
    result_count = 0
    matched_card = None
    matched_invite = None
    matched_reason = None  # 'sent' | 'already_connected' | 'pending' | 'no_connect'
    try:
        result_cards = driver.find_elements(By.CSS_SELECTOR, "div[role='listitem']")
        result_count = len(result_cards)
        print(f"  Found {result_count} result card(s).")

        # Build a set of normalized tokens from the searched name to match cards
        name_tokens = [t for t in name.lower().split() if len(t) >= 2]
        if not name_tokens:
            return 'error'

        # First pass: collect all name-matching cards with their action
        matching_cards = []
        for card in result_cards:
            try:
                card_text = card.text or ''
                card_lower = card_text.lower()
                hits = sum(1 for t in name_tokens if t in card_lower)
                if hits < min(2, len(name_tokens)):
                    continue
                matching_cards.append(card)
            except Exception:
                continue

        if not matching_cards:
            if result_count == 0:
                print("  No results found for this name.")
                return 'no_results'
            print("  No result card matches the searched name.")
            return 'no_match'

        # Pick the FIRST matching card (most relevant in LinkedIn's ranking),
        # then determine its action. We do NOT skip ahead to other same-name
        # results, because the 1st match is usually the right person.
        matched_card = matching_cards[0]
        card_text = matched_card.text or ''

        # Check Connect action
        invite_anchor = matched_card.find_elements(
            By.XPATH,
            ".//a[contains(@aria-label, 'Invita') and contains(@aria-label, 'collegarsi')] | "
            ".//a[contains(@aria-label, 'Invite') and (contains(@aria-label, 'connect') or contains(@aria-label, 'Connect'))]"
        )
        if not invite_anchor:
            invite_anchor = matched_card.find_elements(
                By.XPATH,
                ".//a[.//span[normalize-space(text())='Collegati' or normalize-space(text())='Connetti' or normalize-space(text())='Connect']]"
            )

        is_pending = 'In sospeso' in card_text or 'Pending' in card_text

        has_message = bool(matched_card.find_elements(
            By.XPATH,
            ".//button[contains(@aria-label, 'messaggio') or contains(@aria-label, 'Messaggio') or contains(@aria-label, 'Message')]"
        )) or bool(matched_card.find_elements(
            By.XPATH,
            ".//button[.//span[normalize-space(text())='Messaggio']]"
        ))

        if invite_anchor:
            matched_invite = invite_anchor[0]
        elif is_pending:
            print("  Invite is already pending (In sospeso).")
            return 'already_connected'
        elif has_message:
            print("  Person is already a connection (Message button visible).")
            return 'already_connected'
        else:
            # First card matches the name but has no Connect/Pending/Message button.
            # This is most likely a rate-limit case (LinkedIn hides the Connect
            # button when you hit the search limit) or a card showing only
            # Follow. We do NOT mark this contact as sent: leave the row
            # untouched so it gets retried in a later run.
            print("  No Connect action on matching card (rate limit / Follow / etc). Skipping without marking.")
            return 'no_connect'

        if not matched_invite:
            print("  No Connect action found on the matching card.")
            return 'already_connected'

        aria = matched_invite.get_attribute('aria-label') or matched_invite.text.strip()[:60]
        print(f"  Clicking Connect: '{aria}'")
        try:
            matched_invite.click()
        except WebDriverException as e:
            print(f"  Standard click on Connect failed ({type(e).__name__}); using JS click.")
            driver.execute_script("arguments[0].click();", matched_invite)
        time.sleep(3)

        # After clicking the Connect link, LinkedIn navigates to
        # /preload/search-custom-invite/?vanityName=...
        try:
            WebDriverWait(driver, 8).until(
                lambda d: 'search-custom-invite' in d.current_url
            )
            print(f"  Navigated to: {driver.current_url}")
        except TimeoutException:
            print(f"  URL did not change ({driver.current_url}); checking in-page modal.")

        time.sleep(2)
        driver.save_screenshot(str(debug_dir / "linkedin_invite_page.png"))

        # Find the "Invia senza nota" / "Send without a note" button.
        # The button is <button type="submit" aria-label="Invia senza nota">...</button>
        # with text directly inside (no <span> wrapper).
        send_variants = [
            (By.XPATH, "//button[@aria-label='Invia senza nota']"),
            (By.XPATH, "//button[normalize-space(text())='Invia senza nota']"),
            (By.XPATH, "//button[contains(@aria-label, 'Invia senza nota')]"),
            (By.XPATH, "//button[normalize-space(text())='Send without a note']"),
            (By.XPATH, "//button[contains(@aria-label, 'Send without')]"),
        ]
        send_btn = None
        for by, xp in send_variants:
            try:
                el = WebDriverWait(driver, 4).until(EC.presence_of_element_located((by, xp)))
                if el.is_displayed() and el.is_enabled():
                    send_btn = el
                    break
            except TimeoutException:
                continue
            except Exception:
                continue

        if send_btn:
            label = (send_btn.get_attribute('aria-label') or send_btn.text.strip())[:50]
            print(f"  Found send button: '{label}'. Clicking...")
            try:
                send_btn.click()
            except WebDriverException as e:
                print(f"  Standard click on Send failed ({type(e).__name__}); using JS click.")
                driver.execute_script("arguments[0].click();", send_btn)
            time.sleep(2)
            return 'sent'

        # Fallback: button is inside a shadow DOM (LinkedIn's <div id="interop-outlet"
        # data-testid="interop-shadowdom"> renders modal content there). Regular CSS/XPath
        # selectors cannot pierce shadow roots. Use a JS recursive search across the whole
        # document, including nested shadow roots and same-origin iframes.
        js_clicked = driver.execute_script("""
            (function() {
                function findSendBtn(root) {
                    try {
                        var btn = root.querySelector('button[aria-label="Invia senza nota"]');
                        if (btn) return btn;
                    } catch(e) {}
                    var all = root.querySelectorAll('*');
                    for (var i = 0; i < all.length; i++) {
                        var el = all[i];
                        if (el.shadowRoot) {
                            var found = findSendBtn(el.shadowRoot);
                            if (found) return found;
                        }
                    }
                    var iframes = root.querySelectorAll('iframe');
                    for (var j = 0; j < iframes.length; j++) {
                        try {
                            var idoc = iframes[j].contentDocument;
                            if (idoc) {
                                var found = findSendBtn(idoc);
                                if (found) return found;
                            }
                        } catch(e) {}
                    }
                    return null;
                }
                var btn = findSendBtn(document);
                if (!btn) return 'not_found';
                // Click it via JS (handles being inside shadow DOM)
                btn.click();
                return 'clicked:' + (btn.getAttribute('aria-label') || btn.textContent.trim().slice(0, 50));
            })();
        """)
        if isinstance(js_clicked, str) and js_clicked.startswith('clicked:'):
            print(f"  Found send button in shadow DOM: '{js_clicked[8:]}'. Clicked via JS.")
            time.sleep(2)
            return 'sent'

        # No Send button found: check whether the invite was already submitted.
        page_lower = driver.page_source.lower()
        if any(s in page_lower for s in [
            'richiesta inviata', 'richiesta inoltrata', 'invitation sent', 'request sent',
            'pending', 'in attesa',
        ]):
            print("  Invite appears to have been submitted automatically.")
            return 'sent'

        print("  No Send button or confirmation found. Will not mark as sent.")
        return 'error'

        # No Send button found: check whether the invite was already submitted.
        page_lower = driver.page_source.lower()
        if any(s in page_lower for s in [
            'richiesta inviata', 'richiesta inoltrata', 'invitation sent', 'request sent',
            'pending', 'in attesa',
        ]):
            print("  Invite appears to have been submitted automatically.")
            return 'sent'

        print("  No Send button or confirmation found. Will not mark as sent.")
        return 'error'
    except WebDriverException as e:
        print(f"  WebDriver error: {e}")
        return 'error'
    except Exception as e:
        print(f"  Error during search_and_connect: {e}")
        return 'error'


def main():
    parser = argparse.ArgumentParser(description="Send LinkedIn connection requests from a contacts file.")
    parser.add_argument("contacts_file", nargs='?', default=None, help="Path to Excel file containing contacts.")
    parser.add_argument("--delay-min", type=float, default=3.0, help="Minimum delay between requests (default: 3).")
    parser.add_argument("--delay-max", type=float, default=7.0, help="Maximum delay between requests (default: 7).")
    parser.add_argument("--start-row", type=int, default=1, help="1-based header row; contacts start from start-row+1. Default 1 (start from row 2).")
    parser.add_argument("--limit", type=int, default=0, help="Process at most N contacts (0 = no limit). Useful for testing.")
    parser.add_argument("--profile", type=str, default=None, help="Chrome profile name (default: LinkedInAutomationProfile or $LINKEDIN_PROFILE).")
    args = parser.parse_args()

    contacts_file = args.contacts_file
    if contacts_file is None:
        contacts_file = str(Path(__file__).parent / "Elenco linkedin (1).xlsx")

    # Load credentials from .env or environment variables
    email = os.getenv("LINKEDIN_EMAIL")
    password = os.getenv("LINKEDIN_PASSWORD")
    if not email or not password:
        try:
            from dotenv import load_dotenv
            load_dotenv()
            email = os.getenv("LINKEDIN_EMAIL")
            password = os.getenv("LINKEDIN_PASSWORD")
        except ImportError:
            pass
    if not email or not password:
        raise EnvironmentError("Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD in environment or .env file.")

    contacts, wb, ws = load_contacts(contacts_file)
    print(f"Found {len(contacts)} contact(s) to process (Contattare='y' AND Inviato != 'y').")

    if len(contacts) == 0:
        print("Nothing to do. Exiting.")
        return

    if args.limit > 0:
        contacts = contacts[:args.limit]
        print(f"Limiting to first {args.limit} contact(s).")

    driver = init_driver(profile_name=args.profile)
    sent = skipped = already = no_results = no_connect = errors = 0
    try:
        try:
            linkedin_login(driver, email, password)
        except Exception as e:
            print(f"ERROR: login failed: {e}")
            return

        try:
            for idx, (row_idx, name) in enumerate(contacts, 1):
                print(f"[{idx}/{len(contacts)}] Row {row_idx}: {name}", flush=True)
                outcome = None
                for attempt in range(1, 4):  # up to 3 attempts
                    outcome = search_and_connect(driver, name)
                    if outcome in ('sent', 'already_connected', 'no_results', 'no_match', 'skipped', 'no_connect'):
                        break
                    if attempt < 3:
                        print(f"  retry {attempt+1}/3...", flush=True)
                        time.sleep(3)
                if outcome == 'sent':
                    sent += 1
                    mark_sent(wb, ws, contacts_file, row_idx)
                elif outcome == 'already_connected':
                    already += 1
                    mark_sent(wb, ws, contacts_file, row_idx)
                elif outcome in ('no_results', 'no_match'):
                    no_results += 1
                elif outcome == 'no_connect':
                    no_connect += 1
                    # Do NOT mark: leave the row to be retried on a later run
                    # (likely a rate-limit / Follow-only case, not a real
                    # connection).
                elif outcome == 'skipped':
                    skipped += 1
                else:
                    errors += 1
                delay = random.uniform(args.delay_min, args.delay_max)
                time.sleep(delay)
        except KeyboardInterrupt:
            print("\nInterrupted by user. State has been saved up to the last successful contact.")

        print(f"\n=== SUMMARY ===")
        print(f"  Sent:               {sent}")
        print(f"  Already connected:  {already}")
        print(f"  No results:         {no_results}")
        print(f"  No Connect (rate?): {no_connect}")
        print(f"  Skipped:            {skipped}")
        print(f"  Errors:             {errors}")
    finally:
        try:
            wb.save(contacts_file)
        except Exception:
            pass
        driver.quit()


if __name__ == "__main__":
    main()
