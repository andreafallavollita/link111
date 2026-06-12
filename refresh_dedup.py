"""Refresh the existing_contacts dedup pool.

Sources:
  1. Elenco linkedin (1).xlsx           -> source='elenco_excel'
  2. contact_status_check.json          -> source='contact_status_check'
  3. connections_list.json              -> source='connections_scrape'
  4. (optional) scrape_connections.py   -> refresh #3 first

Idempotent: re-running merges new entries without duplicating.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

import growth_db as db

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "Elenco linkedin (1).xlsx"
STATUS_CHECK_FILE = SCRIPT_DIR / "contact_status_check.json"
CONNECTIONS_FILE = SCRIPT_DIR / "connections_list.json"
SCRAPE_SCRIPT = SCRIPT_DIR / "scrape_connections.py"
PYTHON_EXE = r"C:\Users\andrea.fallavollita\AppData\Local\Programs\Python\Python312\python.exe"


def import_excel(conn, path: Path) -> tuple[int, int]:
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0, 0
    wb = load_workbook(path, data_only=True)
    if "Contatti" not in wb.sheetnames:
        print(f"  [warn] Sheet 'Contatti' missing in {path.name}")
        return 0, 0
    ws = wb["Contatti"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    new_count = 0
    skip_count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if not name:
            continue
        company = str(row[1]).strip() if len(row) > 1 and row[1] else None
        inviato = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ""
        status = "sent" if inviato == "y" else "to_contact"
        before = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        db.upsert_existing_contact(
            conn,
            full_name=name,
            source="elenco_excel",
            status=status,
            company=company,
        )
        after = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        if after > before:
            new_count += 1
        else:
            skip_count += 1
    print(f"  Excel: +{new_count} new, {skip_count} existing/updated")
    return new_count, skip_count


def import_status_check(conn, path: Path) -> tuple[int, int]:
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0, 0
    with path.open("r", encoding="utf-8") as fh:
        records = json.load(fh)
    new_count = 0
    skip_count = 0
    for rec in records:
        name = (rec.get("name") or "").strip()
        if not name:
            continue
        profile_url = rec.get("profile_url") or None
        status = rec.get("status") or None
        before = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        db.upsert_existing_contact(
            conn,
            full_name=name,
            source="contact_status_check",
            profile_url=profile_url,
            status=status,
        )
        after = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        if after > before:
            new_count += 1
        else:
            skip_count += 1
    print(f"  Status check: +{new_count} new, {skip_count} existing/updated")
    return new_count, skip_count


def import_connections_list(conn, path: Path) -> tuple[int, int]:
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0, 0
    with path.open("r", encoding="utf-8") as fh:
        names = json.load(fh)
    new_count = 0
    skip_count = 0
    for entry in names:
        if isinstance(entry, str):
            name = entry.strip()
            profile_url = None
            headline = None
        elif isinstance(entry, dict):
            name = (entry.get("name") or "").strip()
            profile_url = entry.get("profile_url") or None
            headline = entry.get("headline") or None
        else:
            continue
        if not name:
            continue
        before = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        db.upsert_existing_contact(
            conn,
            full_name=name,
            source="connections_scrape",
            profile_url=profile_url,
            status="connected",
            headline=headline,
        )
        after = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts").fetchone()["c"]
        if after > before:
            new_count += 1
        else:
            skip_count += 1
    print(f"  Connections list: +{new_count} new, {skip_count} existing/updated")
    return new_count, skip_count


def run_scrape() -> bool:
    if not SCRAPE_SCRIPT.exists():
        print(f"  [skip] {SCRAPE_SCRIPT.name} not found; cannot refresh connections")
        return False
    print(f"  Launching {SCRAPE_SCRIPT.name} (browser will open)...")
    try:
        result = subprocess.run(
            [PYTHON_EXE, str(SCRAPE_SCRIPT)],
            cwd=str(SCRIPT_DIR),
            timeout=1800,
        )
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        print("  [error] scrape timed out after 30 minutes")
        return False
    except Exception as exc:
        print(f"  [error] scrape failed: {exc}")
        return False


import csv

def import_apollo_csv(conn, path: Path) -> tuple[int, int]:
    if not path.exists():
        return 0, 0
    new_count = 0
    skip_count = 0
    with path.open('r', encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            name = row.get('full_name', '').strip()
            if not name:
                continue
            # check duplicate
            before = conn.execute("SELECT COUNT(*) AS c FROM existing_contacts WHERE full_name = ?", (name,)).fetchone()['c']
            if before:
                skip_count += 1
                continue
            source = row.get('source', 'apollo')
            db.upsert_existing_contact(
                conn,
                full_name=name,
                source=source,
                profile_url=row.get('profile_url'),
                status='discovered',
            )
            new_count += 1
    return new_count, skip_count

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--scrape",
        action="store_true",
        help="Run scrape_connections.py before importing (refreshes connections_list.json)",
    )
    parser.add_argument(
        "--skip-excel",
        action="store_true",
        help="Skip importing the Elenco linkedin xlsx",
    )
    parser.add_argument(
        "--skip-status",
        action="store_true",
        help="Skip importing contact_status_check.json",
    )
    parser.add_argument(
        "--skip-connections",
        action="store_true",
        help="Skip importing connections_list.json",
    )
    parser.add_argument(
        "--import-apollo",
        nargs='*',
        metavar='CSV',
        help="Import one or two Apollo CSV files (clean import and/or linkedin links)",
    )
    args = parser.parse_args()

    started = db.now_iso()
    db.init_db()

    if args.scrape:
        print("[1/4] Refreshing connections via browser scrape...")
        ok = run_scrape()
        if not ok:
            print("[1/4] Scrape failed; will still import existing connections_list.json if present")
    else:
        print("[1/4] Skipping browser scrape (use --scrape to enable)")

    new_total = 0
    skip_total = 0

    with db.db_session() as conn:
        if not args.skip_excel:
            print("[2/4] Importing Elenco linkedin xlsx...")
            n, s = import_excel(conn, EXCEL_FILE)
            new_total += n
            skip_total += s

        if not args.skip_status:
            print("[3/4] Importing contact_status_check.json...")
            n, s = import_status_check(conn, STATUS_CHECK_FILE)
            new_total += n
            skip_total += s

        if not args.skip_connections:
            print("[4/4] Importing connections_list.json...")
            n, s = import_connections_list(conn, CONNECTIONS_FILE)
            new_total += n
            skip_total += s
        if args.import_apollo:
            clean_csv, links_csv = args.import_apollo
            print(f"[5/5] Importing Apollo CSVs: {clean_csv}, {links_csv}...")
            n1, s1 = import_apollo_csv(conn, Path(clean_csv))
            n2, s2 = import_apollo_csv(conn, Path(links_csv))
            new_total += n1 + n2
            skip_total += s1 + s2

        stats = db.stats_overview(conn)
        db.log_run_health(
            conn,
            component="refresh_dedup",
            started_at=started,
            status="ok",
            items_processed=new_total,
            extra={"new": new_total, "skipped": skip_total, "scraped": args.scrape},
        )

    print()
    print("=" * 60)
    print("DEDUP POOL UPDATED")
    print("=" * 60)
    print(f"  New contacts added: {new_total}")
    print(f"  Existing skipped:   {skip_total}")
    print(f"  Total in pool:      {stats['existing_contacts_total']}")
    print(f"  By source:")
    for src, c in stats["existing_by_source"].items():
        print(f"    - {src}: {c}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
