"""Quick accurate check: for each of 187 y contacts, visit the LinkedIn profile
and report the actual state (Message = connesso, In sospeso = pending, Connect =
mai inviata, Segui = follow only).

Does NOT mark Excel. Outputs:
- contact_status_check.csv (full report)
- contact_status_check.json (machine-readable)
- Updates Excel column G "Stato" with one of: connected, pending, connect, follow, no_match
"""

import os
import re
import sys
import time
import json
import random
import argparse
import urllib.parse
import builtins
from datetime import datetime
from pathlib import Path

# Force utf-8 stdout
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass


def _safe_str(s) -> str:
    if s is None:
        return ''
    txt = str(s)
    enc = sys.stdout.encoding or 'utf-8'
    try:
        txt.encode(enc)
        return txt
    except (UnicodeEncodeError, LookupError):
        return txt.encode(enc, errors='replace').decode(enc)


_original_print = builtins.print


def print(*args, **kwargs):
    safe_args = [_safe_str(a) for a in args]
    sep = kwargs.get('sep', ' ')
    if not isinstance(sep, str):
        sep = _safe_str(sep)
    end = kwargs.get('end', '\n')
    if not isinstance(end, str):
        end = _safe_str(end)
    try:
        _original_print(*safe_args, sep=sep, end=end, flush=True)
    except (UnicodeEncodeError, ValueError, OSError):
        try:
            _original_print(*safe_args, sep=sep, end=end)
        except Exception:
            pass


from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "Elenco linkedin (1).xlsx"
SHEET_NAME = "Contatti"
OUT_CSV = SCRIPT_DIR / "contact_status_check.csv"
OUT_JSON = SCRIPT_DIR / "contact_status_check.json"


def init_driver(profile_name: str) -> webdriver.Chrome:
    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36")
    profile_dir = Path(os.environ.get("LOCALAPPDATA", "C:\\Temp")) / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    chrome_options.add_argument(f"--user-data-dir={profile_dir}")
    chrome_options.add_argument("--profile-directory=Default")
    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    driver.implicitly_wait(5)
    return driver


def linkedin_login(driver: webdriver.Chrome, email: str, password: str) -> None:
    driver.get("https://www.linkedin.com/feed")
    time.sleep(5)
    url_lower = driver.current_url.lower()
    if any(x in url_lower for x in ['/login', '/signup', '/checkpoint', '/uas/']):
        print("Not logged in. Going to login page...")
        driver.get("https://www.linkedin.com/login")
        time.sleep(5)
        try:
            email_input = next((e for e in driver.find_elements(By.CSS_SELECTOR, "input[type='email']") if e.is_displayed()), None)
            pwd_input = next((p for p in driver.find_elements(By.CSS_SELECTOR, "input[type='password']") if p.is_displayed()), None)
            if not email_input or not pwd_input:
                raise RuntimeError("Could not find email/password fields.")
            email_input.send_keys(email)
            pwd_input.send_keys(password)
            clicked = False
            for btn in driver.find_elements(By.TAG_NAME, "button"):
                if btn.is_displayed() and btn.text.strip() in ["Accedi", "Sign in", "Log in"]:
                    btn.click()
                    clicked = True
                    break
            if not clicked:
                pwd_input.submit()
        except Exception as e:
            print(f"Login error: {e}")
            raise

    print("Waiting for login (up to 10 min)...")
    deadline = time.time() + 600
    while time.time() < deadline:
        url = driver.current_url.lower()
        if any(x in url for x in ['/login', '/signup', '/uas/']):
            print("Still on login page...")
            time.sleep(3)
            continue
        if '/checkpoint/' in url:
            print("\n*** CHALLENGE - please solve in browser ***")
            time.sleep(10)
            continue
        if '/feed' in url or '/mynetwork' in url:
            print("Login confirmed.")
            return
        try:
            nav = driver.find_element(By.ID, "global-nav")
            if nav and nav.is_displayed():
                print("Login confirmed (#global-nav).")
                return
        except Exception:
            pass
        time.sleep(3)
    raise TimeoutError("Login did not complete.")


def name_tokens(name: str) -> list[str]:
    n = name.strip().lower()
    n = (n.replace('à', 'a').replace('è', 'e').replace('é', 'e')
           .replace('ì', 'i').replace('ò', 'o').replace('ù', 'u'))
    return [t for t in re.split(r'\s+', n) if len(t) >= 2]


def load_contacts(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]
    contacts = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        contattare = ws.cell(row=r, column=4).value
        if name is None:
            continue
        if str(contattare or '').strip().lower() != 'y':
            continue
        contacts.append((r, str(name).strip()))
    return contacts, wb, ws


def find_first_matching_profile_url(driver: webdriver.Chrome, name: str, max_wait: int = 8) -> str | None:
    tokens = name_tokens(name)
    if not tokens:
        return None
    search_url = f"https://www.linkedin.com/search/results/people/?keywords={urllib.parse.quote_plus(name)}"
    driver.get(search_url)
    time.sleep(2)
    try:
        WebDriverWait(driver, max_wait).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='linkedin.com/in/']"))
        )
    except TimeoutException:
        return None
    anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='linkedin.com/in/']")
    for a in anchors:
        try:
            href = a.get_attribute('href') or ''
            if '/in/' not in href:
                continue
            text = (a.text or '').strip()
            if not text:
                continue
            anchor_tokens = name_tokens(text)
            if not anchor_tokens:
                continue
            if set(tokens).issubset(set(anchor_tokens)) or set(anchor_tokens).issubset(set(tokens)):
                return href
        except Exception:
            continue
    return None


def check_action_on_profile(driver: webdriver.Chrome) -> str:
    """Returns: 'connect', 'pending', 'message', 'follow', 'none'."""
    time.sleep(1)
    page_text = (driver.page_source or '').lower()

    # Check Message (means connected)
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            txt = (btn.text or '').strip().lower()
            if 'invia un messaggio' in aria or 'message' in aria or txt == 'messaggio':
                return 'message'
    except Exception:
        pass

    # Check Pending (In sospeso)
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            txt = (btn.text or '').strip().lower()
            if 'in sospeso' in txt or 'pending' in txt or 'in sospeso' in aria:
                return 'pending'
    except Exception:
        pass

    # Check Connect
    try:
        for a in driver.find_elements(By.CSS_SELECTOR, "a[aria-label*='Invita'][aria-label*='collegarsi']"):
            if a.is_displayed():
                return 'connect'
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            if 'invita' in aria and 'collegarsi' in aria:
                return 'connect'
    except Exception:
        pass

    # Check Follow
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            if aria.startswith('segui ') or aria == 'segui' or 'segui ' in aria:
                return 'follow'
    except Exception:
        pass

    return 'none'


def check_contact(driver: webdriver.Chrome, name: str) -> tuple[str, str | None]:
    """Returns (status, profile_url). status in:
    connected, pending, connect, follow, no_match, no_action, error
    """
    try:
        profile_url = find_first_matching_profile_url(driver, name)
    except Exception as e:
        return 'error', None
    if not profile_url:
        return 'no_match', None
    try:
        driver.get(profile_url)
        time.sleep(2)
        action = check_action_on_profile(driver)
    except Exception as e:
        return 'error', profile_url
    if action == 'message':
        return 'connected', profile_url
    if action == 'pending':
        return 'pending', profile_url
    if action == 'connect':
        return 'connect', profile_url
    if action == 'follow':
        return 'follow', profile_url
    return 'no_action', profile_url


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=str(XLSX_PATH))
    parser.add_argument("--profile", default="LinkedInAutomationProfileCheck")
    parser.add_argument("--delay-min", type=float, default=1.5)
    parser.add_argument("--delay-max", type=float, default=3.0)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--start-row", type=int, default=0, help="Resume from this Excel row (1-based, default 0 = from start)")
    parser.add_argument("--no-excel", action="store_true", help="Don't update Excel, only CSV/JSON")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    email = os.getenv("LINKEDIN_EMAIL")
    password = os.getenv("LINKEDIN_PASSWORD")
    if not email or not password:
        try:
            from dotenv import load_dotenv
            load_dotenv()
            email = os.getenv("LINKEDIN_EMAIL")
            password = os.getenv("LINKEDIN_PASSWORD")
        except ImportError:
            pass
    if not email or not password:
        raise EnvironmentError("Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD in environment or .env file.")

    contacts, wb, ws = load_contacts(xlsx_path)
    if args.start_row > 0:
        contacts = [c for c in contacts if c[0] >= args.start_row]
        print(f"Resuming from row {args.start_row}: {len(contacts)} contact(s).")
    else:
        print(f"Found {len(contacts)} contact(s) to check.")

    # Ensure column G "Stato" exists
    if not args.no_excel:
        header_g = ws.cell(row=1, column=7).value
        if header_g != 'Stato':
            ws.cell(row=1, column=7, value='Stato')
            ws.cell(row=1, column=8, value='Stato il')
            try:
                wb.save(xlsx_path)
            except Exception as e:
                print(f"  WARNING: could not save header ({e})")

    if args.limit > 0:
        contacts = contacts[:args.limit]
        print(f"Limiting to {args.limit}.")

    driver = init_driver(args.profile)
    results = []
    counts = {'connected': 0, 'pending': 0, 'connect': 0, 'follow': 0,
              'no_match': 0, 'no_action': 0, 'error': 0}
    try:
        try:
            linkedin_login(driver, email, password)
        except Exception as e:
            print(f"ERROR: login failed: {e}")
            return

        try:
            for idx, (row_idx, name) in enumerate(contacts, 1):
                print(f"[{idx}/{len(contacts)}] Row {row_idx}: {name}", flush=True)
                status, profile_url = check_contact(driver, name)
                counts[status] = counts.get(status, 0) + 1
                print(f"  -> {status}" + (f" ({profile_url})" if profile_url else ""), flush=True)
                results.append({
                    'row': row_idx, 'name': name, 'status': status,
                    'profile_url': profile_url,
                    'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                })
                # Update Excel column G
                if not args.no_excel:
                    ws.cell(row=row_idx, column=7, value=status)
                    ws.cell(row=row_idx, column=8, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    if idx % 5 == 0:
                        try:
                            wb.save(xlsx_path)
                        except Exception as e:
                            print(f"  WARNING: save failed ({e})", flush=True)
                delay = random.uniform(args.delay_min, args.delay_max)
                time.sleep(delay)
        except KeyboardInterrupt:
            print("\nInterrupted.")

        # Final save
        if not args.no_excel:
            try:
                wb.save(xlsx_path)
                print("Excel saved.")
            except Exception as e:
                print(f"WARNING: final save failed ({e})")

        # Save CSV/JSON
        import csv
        with open(OUT_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['row', 'name', 'status', 'profile_url', 'checked_at'])
            writer.writeheader()
            for r in results:
                writer.writerow(r)
        with open(OUT_JSON, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"\n=== SUMMARY ===")
        for k, v in counts.items():
            print(f"  {k:15s} {v:3d}")
        print(f"\nCSV:  {OUT_CSV}")
        print(f"JSON: {OUT_JSON}")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
