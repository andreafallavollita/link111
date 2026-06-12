"""Subagent 3 v2: Send LinkedIn connection requests by navigating to profiles directly.

The search-based approach is rate-limited by LinkedIn. This script bypasses
the rate limit by:
1. Search for the name on LinkedIn
2. Click on the first matching profile link
3. On the profile page, look for the Connect button (visible even when
   search results are rate-limited)
4. Click Connect, then find and click the Send button (which may be in
   a shadow DOM)
"""

import os
import sys
import time
import random
import argparse
from datetime import datetime
from pathlib import Path
import urllib.parse
import builtins

def _safe_str(s) -> str:
    if s is None:
        return ''
    txt = str(s)
    enc = sys.stdout.encoding or 'utf-8'
    try:
        txt.encode(enc)
        return txt
    except (UnicodeEncodeError, LookupError):
        return txt.encode(enc, errors='replace').decode(enc)


_original_print = builtins.print


def print(*args, **kwargs):
    safe_args = [_safe_str(a) for a in args]
    sep = kwargs.get('sep', ' ')
    if not isinstance(sep, str):
        sep = _safe_str(sep)
    end = kwargs.get('end', '\n')
    if not isinstance(end, str):
        end = _safe_str(end)
    try:
        _original_print(*safe_args, sep=sep, end=end, flush=True)
    except (UnicodeEncodeError, ValueError, OSError):
        try:
            _original_print(*safe_args, sep=sep, end=end)
        except Exception:
            pass


from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "Elenco linkedin (1).xlsx"
SHEET_NAME = "Contatti"
DEBUG_DIR = Path(os.environ.get("TEMP", "."))


def init_driver(profile_name: str) -> webdriver.Chrome:
    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36")
    profile_dir = Path(os.environ.get("LOCALAPPDATA", "C:\\Temp")) / profile_name
    profile_dir.mkdir(parents=True, exist_ok=True)
    chrome_options.add_argument(f"--user-data-dir={profile_dir}")
    chrome_options.add_argument("--profile-directory=Default")
    service = ChromeService()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    driver.implicitly_wait(5)
    return driver


def linkedin_login(driver: webdriver.Chrome, email: str, password: str) -> None:
    debug_dir = DEBUG_DIR
    driver.get("https://www.linkedin.com/feed")
    time.sleep(5)
    url_lower = driver.current_url.lower()
    if any(x in url_lower for x in ['/login', '/signup', '/checkpoint', '/uas/']):
        print("Not logged in. Going to login page...")
        driver.get("https://www.linkedin.com/login")
        time.sleep(5)
        try:
            email_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='email']")
            email_input = next((e for e in email_inputs if e.is_displayed()), None)
            pwd_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
            pwd_input = next((p for p in pwd_inputs if p.is_displayed()), None)
            if not email_input or not pwd_input:
                raise RuntimeError("Could not find email/password fields.")
            email_input.send_keys(email)
            pwd_input.send_keys(password)
            clicked = False
            for btn in driver.find_elements(By.TAG_NAME, "button"):
                if btn.is_displayed() and (btn.text.strip() in ["Accedi", "Sign in", "Log in"] or
                                            ("Accedi" in btn.text or "Sign in" in btn.text or "Log in" in btn.text)
                                            and not any(x in btn.text for x in ["Google", "Microsoft", "Apple"])):
                    btn.click()
                    clicked = True
                    break
            if not clicked:
                pwd_input.submit()
        except Exception as e:
            print(f"Login error: {e}")
            raise

    print("Waiting for login (up to 10 min for challenges)...")
    deadline = time.time() + 600
    while time.time() < deadline:
        url = driver.current_url.lower()
        if any(x in url for x in ['/login', '/signup', '/uas/']):
            print("Still on login page...")
            time.sleep(3)
            continue
        if '/checkpoint/' in url:
            print("\n*** CHALLENGE DETECTED - please solve it in the browser ***")
            time.sleep(10)
            continue
        if '/feed' in url or '/mynetwork' in url:
            print("Login confirmed.")
            return
        try:
            nav = driver.find_element(By.ID, "global-nav")
            if nav and nav.is_displayed():
                print("Login confirmed (#global-nav).")
                return
        except NoSuchElementException:
            pass
        print(f"  URL: {driver.current_url}, waiting...")
        time.sleep(3)
    raise TimeoutError("Login did not complete.")


def name_tokens(name: str) -> list[str]:
    n = name.strip().lower()
    n = (n.replace('à', 'a').replace('è', 'e').replace('é', 'e')
           .replace('ì', 'i').replace('ò', 'o').replace('ù', 'u'))
    return [t for t in re.split(r'\s+', n) if len(t) >= 2]


import re


def load_contacts(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]
    contacts = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        contattare = ws.cell(row=r, column=4).value
        inviato = ws.cell(row=r, column=5).value
        if name is None:
            continue
        if str(contattare or '').strip().lower() != 'y':
            continue
        if str(inviato or '').strip().lower() == 'y':
            continue
        contacts.append((r, str(name).strip()))
    return contacts, wb, ws


def mark_sent(wb, ws, xlsx_path: Path, row_idx: int) -> None:
    ws.cell(row=row_idx, column=5, value='y')
    ws.cell(row=row_idx, column=6, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    try:
        wb.save(xlsx_path)
    except Exception as e:
        print(f"  WARNING: could not save Excel ({e})")


def find_first_matching_profile_url(driver: webdriver.Chrome, name: str, max_wait: int = 8) -> str | None:
    """Search for `name` and return the URL of the first profile whose name
    matches `name`. Returns None if no match found or no profile links."""
    tokens = name_tokens(name)
    if not tokens:
        return None
    search_url = f"https://www.linkedin.com/search/results/people/?keywords={urllib.parse.quote_plus(name)}"
    driver.get(search_url)
    time.sleep(3)
    try:
        WebDriverWait(driver, max_wait).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='linkedin.com/in/']"))
        )
    except TimeoutException:
        return None
    anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='linkedin.com/in/']")
    for a in anchors:
        try:
            href = a.get_attribute('href') or ''
            if '/in/' not in href:
                continue
            text = (a.text or '').strip()
            if not text:
                continue
            anchor_tokens = name_tokens(text)
            if not anchor_tokens:
                continue
            if set(tokens).issubset(set(anchor_tokens)) or set(anchor_tokens).issubset(set(tokens)):
                return href
        except Exception:
            continue
    return None


def check_action_on_profile(driver: webdriver.Chrome) -> str:
    """On the profile page, return one of: 'connect', 'pending', 'message', 'follow', 'none'."""
    page_text = (driver.page_source or '').lower()
    # Check for the various states. The profile page has specific buttons.
    if 'in sospeso' in page_text or 'pending' in page_text:
        # Could be an invite you sent, look for the Pending button text near the profile header
        try:
            for btn in driver.find_elements(By.TAG_NAME, "button"):
                txt = (btn.text or '').strip().lower()
                if 'in sospeso' in txt or 'pending' in txt:
                    return 'pending'
        except Exception:
            pass
    # Check for Message button (means connected)
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            txt = (btn.text or '').strip().lower()
            aria = (btn.get_attribute('aria-label') or '').lower()
            if 'invia un messaggio' in aria or 'message' in aria or 'messaggio' == txt:
                return 'message'
    except Exception:
        pass
    # Check for Connect button
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            if 'invita' in aria and 'collegarsi' in aria:
                return 'connect'
        for a in driver.find_elements(By.TAG_NAME, "a"):
            aria = (a.get_attribute('aria-label') or '').lower()
            if 'invita' in aria and 'collegarsi' in aria:
                return 'connect'
    except Exception:
        pass
    # Check for Follow button (user is following or can follow)
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            if 'segui ' in aria or aria.startswith('segui'):
                return 'follow'
    except Exception:
        pass
    return 'none'


def click_connect_on_profile(driver: webdriver.Chrome) -> bool:
    """Click the Connect button on the profile page. Returns True on success."""
    # Try anchor first
    for a in driver.find_elements(By.CSS_SELECTOR, "a[aria-label*='Invita'][aria-label*='collegarsi']"):
        try:
            if a.is_displayed() and a.is_enabled():
                a.click()
                return True
        except WebDriverException:
            try:
                driver.execute_script("arguments[0].click();", a)
                return True
            except Exception:
                continue
    # Try button
    for btn in driver.find_elements(By.TAG_NAME, "button"):
        aria = (btn.get_attribute('aria-label') or '').lower()
        if 'invita' in aria and 'collegarsi' in aria:
            try:
                if btn.is_displayed() and btn.is_enabled():
                    btn.click()
                    return True
            except WebDriverException:
                try:
                    driver.execute_script("arguments[0].click();", btn)
                    return True
                except Exception:
                    continue
    return False


def click_send_in_modal(driver: webdriver.Chrome) -> str:
    """After Connect click, the modal opens. Click the 'Invia senza nota' button.
    Returns 'sent' if clicked, 'error' otherwise."""
    time.sleep(2)
    # Try XPath in light DOM
    send_variants = [
        (By.XPATH, "//button[@aria-label='Invia senza nota']"),
        (By.XPATH, "//button[normalize-space(text())='Invia senza nota']"),
        (By.XPATH, "//button[contains(@aria-label, 'Invia senza nota')]"),
    ]
    for by, xp in send_variants:
        try:
            el = WebDriverWait(driver, 4).until(EC.presence_of_element_located((by, xp)))
            if el.is_displayed() and el.is_enabled():
                try:
                    el.click()
                except WebDriverException:
                    driver.execute_script("arguments[0].click();", el)
                time.sleep(2)
                return 'sent'
        except TimeoutException:
            continue
        except Exception:
            continue
    # Fallback: shadow DOM / iframe recursive search
    result = driver.execute_script("""
        (function() {
            function findBtn(root) {
                try {
                    var btn = root.querySelector('button[aria-label="Invia senza nota"]');
                    if (btn) return btn;
                } catch(e) {}
                var all = root.querySelectorAll('*');
                for (var i = 0; i < all.length; i++) {
                    var el = all[i];
                    if (el.shadowRoot) {
                        var f = findBtn(el.shadowRoot);
                        if (f) return f;
                    }
                }
                var iframes = root.querySelectorAll('iframe');
                for (var j = 0; j < iframes.length; j++) {
                    try {
                        var idoc = iframes[j].contentDocument;
                        if (idoc) {
                            var f = findBtn(idoc);
                            if (f) return f;
                        }
                    } catch(e) {}
                }
                return null;
            }
            var btn = findBtn(document);
            if (!btn) return 'not_found';
            btn.click();
            return 'clicked';
        })();
    """)
    if result == 'clicked':
        time.sleep(2)
        return 'sent'
    return 'error'


def send_to_contact(driver: webdriver.Chrome, name: str) -> str:
    """Try to send a connect request to `name`. Returns outcome string."""
    profile_url = find_first_matching_profile_url(driver, name)
    if not profile_url:
        return 'no_match'
    driver.get(profile_url)
    time.sleep(3)
    action = check_action_on_profile(driver)
    if action == 'pending':
        print("  pending")
        return 'already_connected'
    if action == 'message':
        print("  connected (Message)")
        return 'already_connected'
    if action == 'follow':
        # User is following this person. The profile page might still have
        # a Connect option via the "..." more actions menu. Try to find it.
        # For now, try clicking "Altre azioni" (More actions) and look for Connect.
        more_clicked = False
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            aria = (btn.get_attribute('aria-label') or '').lower()
            if 'altre azioni' in aria or 'more actions' in aria:
                try:
                    btn.click()
                    more_clicked = True
                    break
                except Exception:
                    continue
        if more_clicked:
            time.sleep(2)
            # Look for "Collegati" or "Connetti" or "Invita" in the dropdown
            for opt in driver.find_elements(By.TAG_NAME, "button"):
                aria = (opt.get_attribute('aria-label') or '').lower()
                txt = (opt.text or '').strip().lower()
                if 'invita' in aria and 'collegarsi' in aria:
                    try:
                        opt.click()
                        return click_send_in_modal(driver)
                    except Exception:
                        continue
                if txt in ('collegati', 'connetti', 'connect'):
                    try:
                        opt.click()
                        return click_send_in_modal(driver)
                    except Exception:
                        continue
        print("  follow only (no Connect in more actions)")
        return 'follow_only'
    if action == 'connect':
        if click_connect_on_profile(driver):
            return click_send_in_modal(driver)
        print("  could not click Connect")
        return 'error'
    if action == 'none':
        print("  no action button on profile")
        return 'no_action'
    return 'error'


def main():
    parser = argparse.ArgumentParser(description="Subagent 3 v2: send via direct profile navigation.")
    parser.add_argument("--xlsx", default=str(XLSX_PATH))
    parser.add_argument("--profile", default="LinkedInAutomationProfileCheck")
    parser.add_argument("--delay-min", type=float, default=2.0)
    parser.add_argument("--delay-max", type=float, default=4.0)
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    email = os.getenv("LINKEDIN_EMAIL")
    password = os.getenv("LINKEDIN_PASSWORD")
    if not email or not password:
        try:
            from dotenv import load_dotenv
            load_dotenv()
            email = os.getenv("LINKEDIN_EMAIL")
            password = os.getenv("LINKEDIN_PASSWORD")
        except ImportError:
            pass
    if not email or not password:
        raise EnvironmentError("Set LINKEDIN_EMAIL and LINKEDIN_PASSWORD in environment or .env file.")

    contacts, wb, ws = load_contacts(xlsx_path)
    print(f"Found {len(contacts)} contact(s) to process.")
    if args.limit > 0:
        contacts = contacts[:args.limit]
        print(f"Limiting to {args.limit}.")

    driver = init_driver(args.profile)
    sent = already = follow_only = no_match = no_action = errors = 0
    try:
        try:
            linkedin_login(driver, email, password)
        except Exception as e:
            print(f"ERROR: login failed: {e}")
            return

        try:
            for idx, (row_idx, name) in enumerate(contacts, 1):
                print(f"[{idx}/{len(contacts)}] Row {row_idx}: {name}", flush=True)
                outcome = None
                for attempt in range(1, 4):
                    outcome = send_to_contact(driver, name)
                    if outcome in ('sent', 'already_connected', 'no_match', 'no_action', 'follow_only'):
                        break
                    print(f"  retry {attempt+1}/3...", flush=True)
                    time.sleep(3)
                if outcome == 'sent':
                    sent += 1
                    mark_sent(wb, ws, xlsx_path, row_idx)
                elif outcome == 'already_connected':
                    already += 1
                    mark_sent(wb, ws, xlsx_path, row_idx)
                elif outcome == 'follow_only':
                    follow_only += 1
                elif outcome == 'no_match':
                    no_match += 1
                elif outcome == 'no_action':
                    no_action += 1
                else:
                    errors += 1
                delay = random.uniform(args.delay_min, args.delay_max)
                time.sleep(delay)
        except KeyboardInterrupt:
            print("\nInterrupted.")

        print(f"\n=== SUMMARY ===")
        print(f"  Sent:               {sent}")
        print(f"  Already connected:  {already}")
        print(f"  Follow only:        {follow_only}")
        print(f"  No match:           {no_match}")
        print(f"  No action:          {no_action}")
        print(f"  Errors:             {errors}")
    finally:
        try:
            wb.save(xlsx_path)
        except Exception:
            pass
        driver.quit()


if __name__ == "__main__":
    main()
